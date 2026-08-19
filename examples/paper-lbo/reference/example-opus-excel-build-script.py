"""Build LBO case study model."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

wb = Workbook()

# No iterative calc needed: FCF is computed algebraically to eliminate the
# average-debt-interest circular reference.

# Styles
FONT = "Arial"
input_font = Font(name=FONT, color="0000FF")   # blue = inputs
formula_font = Font(name=FONT, color="000000")  # black = formulas
link_font = Font(name=FONT, color="008000")     # green = cross-sheet
header_font = Font(name=FONT, bold=True, color="000000")
title_font = Font(name=FONT, bold=True, size=12)
default_font = Font(name=FONT)

FMT_CUR = '$#,##0.0;($#,##0.0);"-"'
FMT_PCT = '0.0%;(0.0%);"-"'
FMT_MULT = '0.00"x"'
FMT_YEAR = '0'

# ---------------------------------------------------------------
# 1. Assumptions
# ---------------------------------------------------------------
ws = wb.active
ws.title = "Assumptions"

def set_cell(sheet, coord, value, font=None, fmt=None, comment=None):
    c = sheet[coord]
    c.value = value
    c.font = font or default_font
    if fmt:
        c.number_format = fmt
    if comment:
        c.comment = Comment(comment, "Model")
    return c

# Column widths
ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 15

set_cell(ws, 'A1', 'Paper LBO Case Study - Assumptions', title_font)

set_cell(ws, 'A3', 'Transaction & Financing', header_font)
set_cell(ws, 'A4', 'Close date')
set_cell(ws, 'B4', '12/31/2022', input_font)
set_cell(ws, 'A5', 'Entry multiple (NTM EBITDA)')
set_cell(ws, 'B5', 5.0, input_font, FMT_MULT)
set_cell(ws, 'A6', 'Debt % of purchase price')
set_cell(ws, 'B6', 0.60, input_font, FMT_PCT)
set_cell(ws, 'A7', 'Equity % of purchase price')
set_cell(ws, 'B7', '=1-B6', formula_font, FMT_PCT)
set_cell(ws, 'A8', 'Interest rate')
set_cell(ws, 'B8', 0.10, input_font, FMT_PCT)

set_cell(ws, 'A10', 'Operating Case', header_font)
set_cell(ws, 'A11', '2023 revenue ($)')
set_cell(ws, 'B11', 100.0, input_font, FMT_CUR)
set_cell(ws, 'A12', 'Annual revenue growth')
set_cell(ws, 'B12', 0.10, input_font, FMT_PCT)
set_cell(ws, 'A13', 'EBITDA margin')
set_cell(ws, 'B13', 0.40, input_font, FMT_PCT)
set_cell(ws, 'A14', 'Annual D&A ($)')
set_cell(ws, 'B14', 20.0, input_font, FMT_CUR)
set_cell(ws, 'A15', 'Tax rate')
set_cell(ws, 'B15', 0.40, input_font, FMT_PCT)
set_cell(ws, 'A16', 'Capex (% of revenue)')
set_cell(ws, 'B16', 0.15, input_font, FMT_PCT)
set_cell(ws, 'A17', 'Increase in operating NWC ($ p.a.)')
set_cell(ws, 'B17', 5.0, input_font, FMT_CUR)

set_cell(ws, 'A19', 'Exit', header_font)
set_cell(ws, 'A20', 'Hold period (years)')
set_cell(ws, 'B20', 5, input_font, FMT_YEAR)
set_cell(ws, 'A21', 'Exit multiple (NTM EBITDA)')
set_cell(ws, 'B21', 5.0, input_font, FMT_MULT)

# Reference notes
set_cell(ws, 'A23', 'Notes', header_font)
set_cell(ws, 'A24', 'Purchase price = entry multiple x NTM EBITDA (2023E EBITDA)')
set_cell(ws, 'A25', 'Exit EV = exit multiple x NTM EBITDA at exit (2028E EBITDA)')
set_cell(ws, 'A26', 'Interest = 10% of average period debt balance (iterative calc enabled)')
set_cell(ws, 'A27', '100% of annual FCF sweeps to pay down debt; no interim distributions')
set_cell(ws, 'A28', 'No transaction fees, no cash on balance sheet, no other debt')

# Named references for readability (use direct cell refs to keep formulas explicit)

# ---------------------------------------------------------------
# 2. Sources & Uses
# ---------------------------------------------------------------
ws2 = wb.create_sheet("S&U")
ws2.column_dimensions['A'].width = 40
ws2.column_dimensions['B'].width = 15

set_cell(ws2, 'A1', 'Sources & Uses', title_font)

# NTM EBITDA = 2023 revenue * EBITDA margin
set_cell(ws2, 'A3', 'NTM EBITDA (2023E)', comment=None)
set_cell(ws2, 'B3', "=Assumptions!B11*Assumptions!B13", link_font, FMT_CUR)

set_cell(ws2, 'A4', 'Entry multiple')
set_cell(ws2, 'B4', "=Assumptions!B5", link_font, FMT_MULT)

set_cell(ws2, 'A5', 'Purchase price / Enterprise value')
set_cell(ws2, 'B5', "=B3*B4", formula_font, FMT_CUR)

set_cell(ws2, 'A7', 'Sources', header_font)
set_cell(ws2, 'A8', 'Debt')
set_cell(ws2, 'B8', "=B5*Assumptions!B6", formula_font, FMT_CUR)
set_cell(ws2, 'A9', 'Sponsor equity')
set_cell(ws2, 'B9', "=B5*Assumptions!B7", formula_font, FMT_CUR)
set_cell(ws2, 'A10', 'Total sources')
set_cell(ws2, 'B10', "=SUM(B8:B9)", Font(name=FONT, bold=True), FMT_CUR)

set_cell(ws2, 'A12', 'Uses', header_font)
set_cell(ws2, 'A13', 'Purchase price')
set_cell(ws2, 'B13', "=B5", formula_font, FMT_CUR)
set_cell(ws2, 'A14', 'Total uses')
set_cell(ws2, 'B14', "=SUM(B13:B13)", Font(name=FONT, bold=True), FMT_CUR)

set_cell(ws2, 'A16', 'Check: Sources - Uses')
set_cell(ws2, 'B16', "=B10-B14", formula_font, FMT_CUR)

# ---------------------------------------------------------------
# 3. Operating Model + FCF
# ---------------------------------------------------------------
ws3 = wb.create_sheet("Operating & FCF")
ws3.column_dimensions['A'].width = 30
for col in range(2, 8):
    ws3.column_dimensions[get_column_letter(col)].width = 13

set_cell(ws3, 'A1', 'Operating Model & Free Cash Flow', title_font)

# Header row: years
years = [2023, 2024, 2025, 2026, 2027]
year_cols = ['B', 'C', 'D', 'E', 'F']
set_cell(ws3, 'A3', 'Year')
for i, y in enumerate(years):
    c = set_cell(ws3, f'{year_cols[i]}3', y, header_font)
    c.number_format = '0'
    c.alignment = Alignment(horizontal='center')

# Revenue
set_cell(ws3, 'A5', 'Revenue')
set_cell(ws3, 'B5', "=Assumptions!B11", link_font, FMT_CUR)
for i in range(1, 5):
    set_cell(ws3, f'{year_cols[i]}5', f"={year_cols[i-1]}5*(1+Assumptions!$B$12)", formula_font, FMT_CUR)

# Revenue growth (check)
set_cell(ws3, 'A6', 'Revenue growth %')
set_cell(ws3, 'B6', 'n/a', default_font)
for i in range(1, 5):
    set_cell(ws3, f'{year_cols[i]}6', f"={year_cols[i]}5/{year_cols[i-1]}5-1", formula_font, FMT_PCT)

# EBITDA
set_cell(ws3, 'A8', 'EBITDA')
for i in range(5):
    set_cell(ws3, f'{year_cols[i]}8', f"={year_cols[i]}5*Assumptions!$B$13", formula_font, FMT_CUR)

set_cell(ws3, 'A9', 'EBITDA margin %')
for i in range(5):
    set_cell(ws3, f'{year_cols[i]}9', f"={year_cols[i]}8/{year_cols[i]}5", formula_font, FMT_PCT)

# D&A (shown as negative)
set_cell(ws3, 'A10', 'Less: D&A')
for i in range(5):
    set_cell(ws3, f'{year_cols[i]}10', f"=-Assumptions!$B$14", formula_font, FMT_CUR)

# EBIT
set_cell(ws3, 'A11', 'EBIT')
for i in range(5):
    set_cell(ws3, f'{year_cols[i]}11', f"={year_cols[i]}8+{year_cols[i]}10", formula_font, FMT_CUR)

# Cash interest (from Debt schedule) — negative
set_cell(ws3, 'A12', 'Less: Cash interest')
for i in range(5):
    set_cell(ws3, f'{year_cols[i]}12', f"=-'Debt'!{year_cols[i]}9", link_font, FMT_CUR)

# EBT
set_cell(ws3, 'A13', 'EBT')
for i in range(5):
    set_cell(ws3, f'{year_cols[i]}13', f"={year_cols[i]}11+{year_cols[i]}12", formula_font, FMT_CUR)

# Taxes (negative when EBT positive)
set_cell(ws3, 'A14', 'Less: Taxes')
for i in range(5):
    set_cell(ws3, f'{year_cols[i]}14', f"=-MAX({year_cols[i]}13,0)*Assumptions!$B$15", formula_font, FMT_CUR)

# Net income
set_cell(ws3, 'A15', 'Net income')
for i in range(5):
    set_cell(ws3, f'{year_cols[i]}15', f"={year_cols[i]}13+{year_cols[i]}14", Font(name=FONT, bold=True), FMT_CUR)

# --- Free cash flow ---
set_cell(ws3, 'A18', 'Free Cash Flow', header_font)

set_cell(ws3, 'A19', 'EBITDA')
for i in range(5):
    set_cell(ws3, f'{year_cols[i]}19', f"={year_cols[i]}8", formula_font, FMT_CUR)

set_cell(ws3, 'A20', 'Less: Cash taxes')
for i in range(5):
    set_cell(ws3, f'{year_cols[i]}20', f"={year_cols[i]}14", formula_font, FMT_CUR)

set_cell(ws3, 'A21', 'Less: Cash interest')
for i in range(5):
    set_cell(ws3, f'{year_cols[i]}21', f"={year_cols[i]}12", formula_font, FMT_CUR)

set_cell(ws3, 'A22', 'Less: Capex')
for i in range(5):
    set_cell(ws3, f'{year_cols[i]}22', f"=-{year_cols[i]}5*Assumptions!$B$16", formula_font, FMT_CUR)

set_cell(ws3, 'A23', 'Less: Increase in NWC')
for i in range(5):
    set_cell(ws3, f'{year_cols[i]}23', f"=-Assumptions!$B$17", formula_font, FMT_CUR)

set_cell(ws3, 'A24', 'Free cash flow')
for i in range(5):
    set_cell(ws3, f'{year_cols[i]}24', f"=SUM({year_cols[i]}19:{year_cols[i]}23)", Font(name=FONT, bold=True), FMT_CUR)

# Reconciliation check: presentation FCF equals algebraic FCF on Debt sheet
set_cell(ws3, 'A25', 'Check: presentation vs. algebraic FCF')
for i in range(5):
    set_cell(ws3, f'{year_cols[i]}25', f"={year_cols[i]}24-Debt!{year_cols[i]}6", formula_font, FMT_CUR)

# ---------------------------------------------------------------
# 4. Debt schedule
# ---------------------------------------------------------------
ws4 = wb.create_sheet("Debt")
ws4.column_dimensions['A'].width = 30
for col in range(2, 8):
    ws4.column_dimensions[get_column_letter(col)].width = 13

set_cell(ws4, 'A1', 'Debt Schedule', title_font)

set_cell(ws4, 'A3', 'Year')
for i, y in enumerate(years):
    c = set_cell(ws4, f'{year_cols[i]}3', y, header_font)
    c.number_format = '0'
    c.alignment = Alignment(horizontal='center')

# Beginning debt
set_cell(ws4, 'A5', 'Beginning debt')
set_cell(ws4, 'B5', "='S&U'!B8", link_font, FMT_CUR)
for i in range(1, 5):
    set_cell(ws4, f'{year_cols[i]}5', f"={year_cols[i-1]}8", formula_font, FMT_CUR)

# FCF (algebraic solution to avoid the average-debt-interest circular reference).
# Derivation: Interest_t = ir * (Beg + End)/2 = ir*Beg - ir/2 * FCF_t
#            Taxes_t    = tax * (EBIT_t - Interest_t)
#            FCF_t      = EBITDA_t - Taxes_t - Interest_t - Capex_t - dNWC
# Solving for FCF_t (with EBIT = EBITDA - DA) yields:
#   FCF_t = ((1-tax)*EBITDA_t + tax*DA - (1-tax)*ir*Beg_t - Capex_t - dNWC)
#           / (1 - (1-tax)*ir*0.5)
set_cell(ws4, 'A6', 'Free cash flow (algebraic)',
         comment="Solved directly from assumptions and beginning debt to avoid the "
                 "circular reference between FCF -> ending debt -> average debt -> "
                 "interest -> taxes -> FCF. Result is mathematically identical to "
                 "the presentation total on the Operating & FCF sheet.")
for i in range(5):
    rev_ref = f"'Operating & FCF'!{year_cols[i]}5"
    beg_ref = f"{year_cols[i]}5"
    numerator = (
        f"((1-Assumptions!$B$15)*{rev_ref}*Assumptions!$B$13"
        f"+Assumptions!$B$15*Assumptions!$B$14"
        f"-(1-Assumptions!$B$15)*Assumptions!$B$8*{beg_ref}"
        f"-{rev_ref}*Assumptions!$B$16"
        f"-Assumptions!$B$17)"
    )
    denominator = "(1-(1-Assumptions!$B$15)*Assumptions!$B$8*0.5)"
    set_cell(ws4, f'{year_cols[i]}6', f"={numerator}/{denominator}", formula_font, FMT_CUR)

# Paydown (negative of FCF)
set_cell(ws4, 'A7', 'Less: FCF sweep (paydown)')
for i in range(5):
    set_cell(ws4, f'{year_cols[i]}7', f"=-{year_cols[i]}6", formula_font, FMT_CUR)

# Ending debt
set_cell(ws4, 'A8', 'Ending debt')
for i in range(5):
    set_cell(ws4, f'{year_cols[i]}8', f"={year_cols[i]}5+{year_cols[i]}7", Font(name=FONT, bold=True), FMT_CUR)

# Interest = 10% * average of beg & end
set_cell(ws4, 'A9', 'Interest expense (avg debt)')
for i in range(5):
    set_cell(ws4, f'{year_cols[i]}9', f"=Assumptions!$B$8*AVERAGE({year_cols[i]}5,{year_cols[i]}8)", formula_font, FMT_CUR)

# Roll check
set_cell(ws4, 'A11', 'Check: Beg + Paydown - End')
for i in range(5):
    set_cell(ws4, f'{year_cols[i]}11', f"={year_cols[i]}5+{year_cols[i]}7-{year_cols[i]}8", formula_font, FMT_CUR)

# ---------------------------------------------------------------
# 5. Returns
# ---------------------------------------------------------------
ws5 = wb.create_sheet("Returns")
ws5.column_dimensions['A'].width = 35
for col in range(2, 8):
    ws5.column_dimensions[get_column_letter(col)].width = 13

set_cell(ws5, 'A1', 'Exit / Returns', title_font)

# Exit EBITDA = NTM at exit = 2028E EBITDA = 2027 revenue * (1+g) * margin
set_cell(ws5, 'A3', 'Exit year')
set_cell(ws5, 'B3', "=2022+Assumptions!B20", formula_font, '0')

set_cell(ws5, 'A4', 'NTM EBITDA at exit (2028E)')
# 2028 revenue = 2027 revenue * (1+g); EBITDA = *margin
set_cell(ws5, 'B4', "='Operating & FCF'!F5*(1+Assumptions!B12)*Assumptions!B13", link_font, FMT_CUR)

set_cell(ws5, 'A5', 'Exit multiple')
set_cell(ws5, 'B5', "=Assumptions!B21", link_font, FMT_MULT)

set_cell(ws5, 'A6', 'Exit enterprise value')
set_cell(ws5, 'B6', "=B4*B5", formula_font, FMT_CUR)

set_cell(ws5, 'A7', 'Less: Ending debt at exit')
set_cell(ws5, 'B7', "=-Debt!F8", link_font, FMT_CUR)

set_cell(ws5, 'A8', 'Exit equity value')
set_cell(ws5, 'B8', "=B6+B7", Font(name=FONT, bold=True), FMT_CUR)

set_cell(ws5, 'A10', 'Sponsor equity invested')
set_cell(ws5, 'B10', "='S&U'!B9", link_font, FMT_CUR)

set_cell(ws5, 'A11', 'MoM')
set_cell(ws5, 'B11', "=B8/B10", formula_font, FMT_MULT)

set_cell(ws5, 'A12', 'IRR')
set_cell(ws5, 'B12', "=(B8/B10)^(1/Assumptions!B20)-1", formula_font, FMT_PCT)

# Equity bridge reconciliation
set_cell(ws5, 'A15', 'Equity bridge check', header_font)
set_cell(ws5, 'A16', 'Exit EV')
set_cell(ws5, 'B16', "=B6", formula_font, FMT_CUR)
set_cell(ws5, 'A17', 'Less: Debt at exit')
set_cell(ws5, 'B17', "=B7", formula_font, FMT_CUR)
set_cell(ws5, 'A18', 'Exit equity (bridge)')
set_cell(ws5, 'B18', "=B16+B17", formula_font, FMT_CUR)
set_cell(ws5, 'A19', 'Check: bridge - direct')
set_cell(ws5, 'B19', "=B18-B8", formula_font, FMT_CUR)

# Debt paydown reconciliation
set_cell(ws5, 'A21', 'Debt paydown reconciliation', header_font)
set_cell(ws5, 'A22', 'Beginning debt at close')
set_cell(ws5, 'B22', "=Debt!B5", link_font, FMT_CUR)
set_cell(ws5, 'A23', 'Total FCF swept (sum of paydowns)')
set_cell(ws5, 'B23', "=-SUM(Debt!B7:F7)", formula_font, FMT_CUR)
set_cell(ws5, 'A24', 'Ending debt at exit')
set_cell(ws5, 'B24', "=B22-B23", formula_font, FMT_CUR)
set_cell(ws5, 'A25', 'Check: reconciliation vs. schedule')
set_cell(ws5, 'B25', "=B24-Debt!F8", formula_font, FMT_CUR)

# ---------------------------------------------------------------
# 6. Sensitivity
# ---------------------------------------------------------------
ws6 = wb.create_sheet("Sensitivity")
ws6.column_dimensions['A'].width = 22
for col in range(2, 20):
    ws6.column_dimensions[get_column_letter(col)].width = 12

set_cell(ws6, 'A1', 'IRR Sensitivity: Exit Multiple x Revenue Growth', title_font)

# Helper table: for each growth rate row, compute Rev1..5, BegDebt/FCF/EndDebt for years 1..5, exit EBITDA
# Layout:
# A: growth rate label
# B: growth rate value (input, blue)
# C-G: Rev year 1..5
# H-L: BegDebt year 1..5
# M-Q: FCF year 1..5
# R-V: EndDebt year 1..5
# W: NTM EBITDA at exit (2028)

set_cell(ws6, 'A3', 'Helper table (per growth rate)', header_font)
set_cell(ws6, 'A4', 'Growth', header_font)
for i, y in enumerate(years):
    set_cell(ws6, f'{get_column_letter(3+i)}4', f'Rev {y}', header_font)
for i, y in enumerate(years):
    set_cell(ws6, f'{get_column_letter(8+i)}4', f'BegDebt {y}', header_font)
for i, y in enumerate(years):
    set_cell(ws6, f'{get_column_letter(13+i)}4', f'FCF {y}', header_font)
for i, y in enumerate(years):
    set_cell(ws6, f'{get_column_letter(18+i)}4', f'EndDebt {y}', header_font)
set_cell(ws6, 'W4', 'NTM EBITDA (2028)', header_font)

# Growth rates
growth_rates = [0.06, 0.08, 0.10, 0.12, 0.14]
for i, g in enumerate(growth_rates):
    row = 5 + i
    set_cell(ws6, f'A{row}', g, input_font, FMT_PCT)
    # Revenue: C=Y1, D=Y2, E=Y3, F=Y4, G=Y5
    set_cell(ws6, f'C{row}', f"=Assumptions!$B$11", formula_font, FMT_CUR)
    for j in range(1, 5):
        prev = get_column_letter(3 + j - 1)
        cur = get_column_letter(3 + j)
        set_cell(ws6, f'{cur}{row}', f"={prev}{row}*(1+$A{row})", formula_font, FMT_CUR)

    # BegDebt: H=Y1 (= initial debt), I..L = prior year EndDebt
    set_cell(ws6, f'H{row}', "='S&U'!$B$8", formula_font, FMT_CUR)
    # For years 2-5, BegDebt = prior EndDebt
    # EndDebt columns: R=Y1, S=Y2, T=Y3, U=Y4, V=Y5
    end_cols = ['R', 'S', 'T', 'U', 'V']
    beg_cols = ['H', 'I', 'J', 'K', 'L']
    fcf_cols = ['M', 'N', 'O', 'P', 'Q']
    rev_cols = ['C', 'D', 'E', 'F', 'G']
    for j in range(1, 5):
        set_cell(ws6, f'{beg_cols[j]}{row}', f"={end_cols[j-1]}{row}", formula_font, FMT_CUR)
    # FCF and EndDebt formulas (algebraic to avoid circularity within sensitivity)
    # FCF_t = ((1-tax)*EBITDA + tax*DA - (1-tax)*ir*BegDebt - Capex - NWC) / (1 - (1-tax)*ir*0.5)
    for j in range(5):
        rev = f'{rev_cols[j]}{row}'
        beg = f'{beg_cols[j]}{row}'
        # numerator uses Assumptions references
        numerator = (
            f"((1-Assumptions!$B$15)*{rev}*Assumptions!$B$13 "
            f"+ Assumptions!$B$15*Assumptions!$B$14 "
            f"- (1-Assumptions!$B$15)*Assumptions!$B$8*{beg} "
            f"- {rev}*Assumptions!$B$16 "
            f"- Assumptions!$B$17)"
        )
        denominator = "(1 - (1-Assumptions!$B$15)*Assumptions!$B$8*0.5)"
        set_cell(ws6, f'{fcf_cols[j]}{row}', f"={numerator}/{denominator}", formula_font, FMT_CUR)
        set_cell(ws6, f'{end_cols[j]}{row}', f"={beg}-{fcf_cols[j]}{row}", formula_font, FMT_CUR)

    # NTM EBITDA at exit = year 5 rev * (1+g) * margin
    set_cell(ws6, f'W{row}', f"=G{row}*(1+$A{row})*Assumptions!$B$13", formula_font, FMT_CUR)

# IRR grid: rows = growth (6% to 14%), columns = exit multiple (3.0x to 7.0x)
grid_top = 13
set_cell(ws6, f'A{grid_top}', 'IRR Sensitivity Grid', header_font)
set_cell(ws6, f'A{grid_top+1}', 'Growth \\ Multiple', header_font)
exit_multiples = [3.0, 4.0, 5.0, 6.0, 7.0]
for j, m in enumerate(exit_multiples):
    c = set_cell(ws6, f'{get_column_letter(2+j)}{grid_top+1}', m, input_font, FMT_MULT)
    c.alignment = Alignment(horizontal='center')

for i, g in enumerate(growth_rates):
    row = grid_top + 2 + i
    helper_row = 5 + i
    set_cell(ws6, f'A{row}', g, input_font, FMT_PCT)
    for j, m in enumerate(exit_multiples):
        col = get_column_letter(2 + j)
        multiple_cell = f'{get_column_letter(2+j)}${grid_top+1}'
        # IRR = ((multiple * NTM EBITDA - EndDebt_Y5) / SponsorEquity)^(1/HoldPeriod) - 1
        formula = (
            f"=({multiple_cell}*W{helper_row}-V{helper_row})"
            f"/'S&U'!$B$9"
        )
        formula = f"=(({multiple_cell}*$W{helper_row}-$V{helper_row})/'S&U'!$B$9)^(1/Assumptions!$B$20)-1"
        set_cell(ws6, f'{col}{row}', formula, formula_font, FMT_PCT)

# Legend
set_cell(ws6, f'A{grid_top+8}', 'Note: sensitivity holds all other assumptions constant. Uses algebraic solution for FCF to avoid iterative-calc dependency.', default_font)

# ---------------------------------------------------------------
# Save
# ---------------------------------------------------------------
wb.save('/app/model.xlsx')
print("Saved /app/model.xlsx")
